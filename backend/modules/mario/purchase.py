import pandas as pd
import numpy as np
import io
import re
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# 1. Smarter Tax Formatting Magic (Fixed 'S' to 'P')
def format_tax_rate(label, account):
    if pd.isna(label): return label
    label_str = str(label)
    account_str = str(account).lower()
    
    # Only trigger if there is a number explicitly followed by a % sign
    match = re.search(r'([0-9.]+)\s*%', label_str)
    if match:
        num = float(match.group(1))
        
        # If it's an IGST transaction, do NOT double it
        if 'igst' in account_str or 'igst' in label_str.lower():
            return f"{num:g}% IGST P"
        
        # If it's CGST or SGST, double it
        doubled = num * 2
        return f"{doubled:g}% GST P"
        
    # If no % sign is found (like bank descriptions), return exactly as is!
    return label_str

def clean_purchase_data(df, purchase_type='Regular'):
    # 0. BULLETPROOFING: Reset row index and drop any duplicate columns immediately
    df = df.reset_index(drop=True)
    df = df.loc[:, ~df.columns.duplicated()].copy()

    # Ensure required columns exist to avoid crashes
    for col in ['Account', 'Label', 'Date', 'Debit', 'Credit', 'Taxable Amt.']:
        if col not in df.columns: df[col] = 0.0 if 'Amt' in col or col in ['Debit', 'Credit'] else ''

    df['Debit'] = pd.to_numeric(df['Debit'], errors='coerce').fillna(0)
    df['Credit'] = pd.to_numeric(df['Credit'], errors='coerce').fillna(0)
    df['Taxable Amt.'] = pd.to_numeric(df['Taxable Amt.'], errors='coerce').fillna(0)

    # 1. IDENTIFY SHEET TABS (Magic sorting for CNs)
    if purchase_type == 'RCM':
        is_cn = df['Debit'] != 0  # In RCM, Debit means Credit Note
        df['Sheet_Category'] = np.where(is_cn, 'CN', 'RCM')
    elif purchase_type == 'Import':
        is_cn = df['Credit'] != 0
        df['Sheet_Category'] = np.where(is_cn, 'CN', 'Import')
    else: # Regular
        is_cn = df['Credit'] != 0
        df['Sheet_Category'] = np.where(is_cn, 'CN', 'Purchase')

    # 2. Extract Tax Amount (Keeping it positive for Reco as requested!)
    tax_amount = df[['Debit', 'Credit']].max(axis=1)

    # Identify tax types from the Account column
    account_series = df['Account'].astype(str)
    is_igst = account_series.str.contains('igst', case=False, na=False)
    is_cgst = account_series.str.contains('cgst', case=False, na=False)
    is_sgst = account_series.str.contains('sgst', case=False, na=False)

    df['IGST'] = np.where(is_igst, tax_amount, 0.0)
    df['CGST'] = np.where(is_cgst, tax_amount, 0.0)
    df['SGST'] = np.where(is_sgst, tax_amount, 0.0)

    # Auto-fill SGST if missing
    mask_missing_sgst = (df['CGST'] != 0) & (df['SGST'] == 0)
    df.loc[mask_missing_sgst, 'SGST'] = df.loc[mask_missing_sgst, 'CGST']

    # Rename Columns
    df = df.rename(columns={
        'Partner': 'Vendor Name',
        'Number': 'Invoice Number',
        'Label': 'Tax Rate',
        'Taxable Amt.': 'Taxable Amount'
    })
    
    # BULLETPROOFING PART 2
    df = df.loc[:, ~df.columns.duplicated()].copy()
    
    # Apply the new smart tax formatter, passing both the Label and the Account Name
    df['Tax Rate'] = df.apply(lambda row: format_tax_rate(row['Tax Rate'], row['Account']), axis=1)
    
    # Rename CGST/SGST to GST in the Account column, leaving IGST alone
    df['Account'] = df['Account'].astype(str).str.replace(r'[CS]GST', 'GST', regex=True, flags=re.IGNORECASE)
    
    # Calculate Total based on the raw positive values
    df['Total'] = df['Taxable Amount'] + df['IGST'] + df['CGST'] + df['SGST']

    if 'Date' in df.columns:
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce').dt.strftime('%Y-%m-%d')

    df['Type'] = purchase_type

    final_columns = [
        'Vendor Name', 'GSTIN', 'Date', 'Reference', 'Invoice Number', 
        'Account', 'Tax Rate', 'Type', 'Total', 'Taxable Amount', 'IGST', 'CGST', 'SGST', 'Sheet_Category'
    ]
    
    for col in final_columns:
        if col not in df.columns: df[col] = '' 
            
    return df[final_columns]

# --- THE MAIN ENGINE ROUTE ---
def generate_mario_purchase_report(files_dict):
    dataframes = []

    def load_and_process(file_key, purchase_type):
        if file_key in files_dict and files_dict[file_key].filename != '':
            df = pd.read_excel(files_dict[file_key])
            cleaned_df = clean_purchase_data(df, purchase_type=purchase_type)
            dataframes.append(cleaned_df)

    # Process files
    load_and_process('file_reg_cgst', 'Regular')
    load_and_process('file_reg_igst', 'Regular')
    load_and_process('file_rcm_cgst', 'RCM')
    load_and_process('file_rcm_igst', 'RCM')
    load_and_process('file_import_cgst', 'Import')
    load_and_process('file_import_igst', 'Import')

    if not dataframes:
        raise ValueError("Please upload at least one valid file.")

    combined_df = pd.concat(dataframes, ignore_index=True)

    # Output Stream
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name in ['Purchase', 'RCM', 'Import', 'CN']:
            sheet_df = combined_df[combined_df['Sheet_Category'] == sheet_name].copy()
            
            if not sheet_df.empty:
                sheet_df = sheet_df.drop(columns=['Sheet_Category'])
                
                # We removed the pandas summary_row here so it doesn't double up!
                sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Reload with openpyxl in memory to add formatting
    output.seek(0)
    wb = openpyxl.load_workbook(output)
    sum_cols = ['Total', 'Taxable Amount', 'IGST', 'CGST', 'SGST']

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row
        if max_row < 2: continue

        ws.auto_filter.ref = ws.dimensions
        
        total_row_idx = max_row + 1
        ws.cell(row=total_row_idx, column=1).value = "GRAND TOTAL"
        ws.cell(row=total_row_idx, column=1).font = Font(bold=True)

        headers = [cell.value for cell in ws[1]]
        for col_name in sum_cols:
            if col_name in headers:
                col_idx = headers.index(col_name) + 1
                col_letter = get_column_letter(col_idx)
                
                formula = f"=SUBTOTAL(9, {col_letter}2:{col_letter}{max_row})"
                
                cell = ws.cell(row=total_row_idx, column=col_idx)
                cell.value = formula
                cell.font = Font(bold=True)
                cell.number_format = '#,##0.00'

    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    return final_output