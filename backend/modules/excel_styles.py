"""Shared openpyxl styling helpers so every generated report reads as one
consistent, professional product instead of each module inventing its own
look. Pure presentation -- no business logic."""
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TOTAL_FONT = Font(bold=True)
THIN_BORDER = Border(bottom=Side(style='thin', color="BFBFBF"))
NEGATIVE_FONT = Font(color="C00000")
NUMBER_FORMAT = '#,##0.00;[Red](#,##0.00)'


def style_header_row(ws, n_cols, row=1):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')


def style_data_rows(ws, first_row, last_row, first_col, last_col, numeric_cols):
    for r in range(first_row, last_row + 1):
        for c in range(first_col, last_col + 1):
            cell = ws.cell(row=r, column=c)
            if c in numeric_cols:
                cell.number_format = NUMBER_FORMAT
                if isinstance(cell.value, (int, float)) and cell.value < 0:
                    cell.font = NEGATIVE_FONT


def autofit_columns(ws, df, min_width=10, max_width=40):
    for i, col in enumerate(df.columns, start=1):
        sample = df[col].astype(str).head(200)
        width = max([len(str(col))] + [len(v) for v in sample]) + 2
        ws.column_dimensions[get_column_letter(i)].width = max(min_width, min(width, max_width))
