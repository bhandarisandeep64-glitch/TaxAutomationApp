import re
import pandas as pd
import os
from modules.excel_styles import style_header_row, style_data_rows, autofit_columns
from modules.direct_tax.tds_section_mapping import lookup_new_section

# Odoo's Label format changed alongside the Income Tax Act 2025 rollout --
# real exports now read "New 194H 2%" (section then rate, "New" prefix,
# no trailing space after the final %), not the older "2% 194C" (rate then
# section) this parser originally assumed. That old form is kept as a
# fallback for historical periods that may still carry it.
_LABEL_NEW_FORMAT_RE = re.compile(r'new\s+(\S+)\s+(\d*\.?\d+)\s*%', re.IGNORECASE)
_LABEL_LEGACY_FORMAT_RE = re.compile(r'(\d*\.?\d+)\s*%\s+(\S+)')

# Odoo's own Account name now embeds the new section reference and TRACES
# code directly, e.g. "112456 TDS 393(1)1(ii) - TDS ON COMMISSION/BROKERAGE
# - 1006" -- self-updating as the client's chart of accounts evolves, so
# it's used as the primary source for New Section/Section Code here (unlike
# Zoho, which has no equivalent field and stays on the hardcoded table).
# Falls back to the shared lookup table when an account name doesn't follow
# this pattern (e.g. older entries, or an account not yet renamed).
_ACCOUNT_CODE_RE = re.compile(r'TDS\s+(.+?)\s*-\s*(.+?)\s*-\s*(\d+)\s*$')
_SECTION_REF_SPACING_RE = re.compile(r'^(\d+\([^)]+\))(\d.*)$')


def _parse_label(label):
    """Returns (old_section, rate) from an Odoo Label string, trying the
    current format first and falling back to the legacy one."""
    text = str(label or '')
    m = _LABEL_NEW_FORMAT_RE.search(text)
    if m:
        return m.group(1), float(m.group(2))
    m = _LABEL_LEGACY_FORMAT_RE.search(text)
    if m:
        return m.group(2), float(m.group(1))
    return 'N/A', None


def _format_section_ref(ref):
    """'393(1)1(ii)' -> '393(1), Sl. 1(ii)', matching the shared lookup
    table's own formatting so a report doesn't mix two different styles
    depending on which source resolved a given row. Falls back to the raw
    text if it doesn't match this two-group shape."""
    m = _SECTION_REF_SPACING_RE.match(ref)
    return f"{m.group(1)}, Sl. {m.group(2)}" if m else ref


def _parse_account_new_section(account):
    """Extracts (new_section, code) directly from Odoo's own Account name.
    Returns (None, None) if the account name doesn't follow the expected
    pattern, so the caller can fall back to the shared lookup table."""
    m = _ACCOUNT_CODE_RE.search(str(account or ''))
    if not m:
        return None, None
    return _format_section_ref(m.group(1).strip()), m.group(3).strip()


def process_tds_odoo(file_paths, output_folder, custom_filename=None):
    """
    Processes Multiple Odoo files.
    - Merges all uploaded files.
    - Appends Summary to the SAME sheet.

    Unlike Zoho (one file = one section, from the filename), Odoo's ledger
    export carries the rate and the old section as text inside a single
    "Label" column per row (e.g. "New 194H 2%") -- so a single upload can
    mix multiple sections, and Old Section/rate are derived per row rather
    than per file. New Section/Section Code are read from Odoo's own
    Account name first (self-updating), falling back to the shared Income
    Tax Act 2025 lookup table (tds_section_mapping.py) when that fails.
    """
    try:
        all_data = []

        # 1. Read All Files
        for fp in file_paths:
            try:
                if fp.endswith('.csv'):
                    df_temp = pd.read_csv(fp)
                else:
                    df_temp = pd.read_excel(fp)
                all_data.append(df_temp)
            except Exception as e:
                return {"success": False, "error": f"Failed to read file {os.path.basename(fp)}: {str(e)}"}

        if not all_data:
            return {"success": False, "error": "No valid data found in uploaded files."}

        # 2. Merge DataFrames
        df = pd.concat(all_data, ignore_index=True)

        if 'Credit' in df.columns:
            df.rename(columns={'Credit': 'TDS'}, inplace=True)

        # --- LOGIC BLOCK ---
        parsed = df['Label'].apply(_parse_label)
        df['Old Section'] = parsed.apply(lambda t: t[0])
        df['Rates'] = parsed.apply(lambda t: t[1])

        def get_pan_type(pan):
            pan_str = str(pan)
            return 'Co.' if len(pan_str) >= 4 and pan_str[3].upper() == 'C' else 'Non Co.'

        df['Co./Non Co.'] = df['PAN No.'].apply(get_pan_type)
        df['Checking for rates'] = df['Taxable Amt.'] * (df['Rates'] / 100)
        df['Difference'] = df['TDS'] - df['Checking for rates']
        df['Remarks'] = df['Difference'].apply(lambda x: 'match' if abs(x) < 1 else 'mismatch')

        rename_map = {
            'Partner': 'Vendor', 'PAN No.': 'Permanent Account Number (PAN)', 'Number': 'Transaction#',
            'Journal': 'Transaction Type', 'Taxable Amt.': 'Total', 'TDS': 'Tax Deducted at Source',
            'Rates': 'Rate at which deducted'
        }
        df.rename(columns=rename_map, inplace=True)

        df['Total After TDS Deduction'] = df['Total'] - df['Tax Deducted at Source']

        # New Section / Section Code -- Odoo's own Account name first
        # (self-updating), the shared lookup table as fallback.
        new_sections, section_codes = [], []
        for _, row in df.iterrows():
            new_sec, code = _parse_account_new_section(row.get('Account'))
            if not new_sec:
                new_sec, code = lookup_new_section(row['Old Section'], row['Rate at which deducted'])
            new_sections.append(new_sec)
            section_codes.append(code)
        df['New Section'] = new_sections
        df['Section Code'] = section_codes

        df['Date_dt'] = pd.to_datetime(df['Date'], errors='coerce')
        valid_dates = df['Date_dt'].dropna()

        report_month_period = None
        if not valid_dates.empty:
             modes = valid_dates.dt.to_period('M').mode()
             if not modes.empty:
                report_month_period = modes[0]

        if report_month_period:
            df['Months Difference'] = df['Date_dt'].dt.to_period('M').apply(
                lambda x: (report_month_period - x).n if pd.notna(x) else 0)
            df['Interest'] = df.apply(
                lambda row: row['Tax Deducted at Source'] * 0.015 * (row['Months Difference'] + 1)
                if row['Months Difference'] > 0 else 0, axis=1).round(0)
        else:
            df['Interest'] = 0

        df['Date'] = df['Date_dt'].dt.strftime('%Y-%m-%d').fillna('')

        # Reorder (blank Challan placeholder columns removed -- handled by
        # the separate TDS Challan Mapper tool instead)
        final_column_order = [
            'Vendor', 'Permanent Account Number (PAN)', 'Date', 'Reference', 'Transaction#',
            'Account', 'Transaction Type', 'Label', 'Total', 'Tax Deducted at Source',
            'Checking for rates', 'Difference', 'Remarks', 'Interest', 'Rate at which deducted',
            'Old Section', 'New Section', 'Section Code', 'Co./Non Co.',
        ]

        existing_cols = [c for c in final_column_order if c in df.columns]
        detailed_df = df[existing_cols].copy()

        # Section Summary
        summary_df = pd.DataFrame()
        if {'Old Section', 'New Section', 'Section Code', 'Co./Non Co.'}.issubset(detailed_df.columns):
            summary_df = pd.pivot_table(detailed_df, index=['Old Section', 'New Section', 'Section Code', 'Co./Non Co.'],
                                    values=['Tax Deducted at Source', 'Interest'],
                                    aggfunc='sum', margins=True, margins_name='Total')
            summary_df.reset_index(inplace=True)
            summary_df.rename(columns={'Tax Deducted at Source': 'Total Tax Deducted', 'Interest': 'Total Interest'}, inplace=True)
            summary_df['Total Tax Deducted'] = summary_df['Total Tax Deducted'].round(0)
            summary_df['Total Interest'] = summary_df['Total Interest'].round(0)

        # Vendor Summary
        if 'Vendor' in detailed_df.columns:
            vendor_summary_df = detailed_df.groupby('Vendor').agg(
                **{'Total Taxable Value': ('Total', 'sum'), 'Total TDS': ('Tax Deducted at Source', 'sum')}
            ).reset_index().rename(columns={'Vendor': 'Vendor Name'})
            vendor_summary_df = vendor_summary_df.sort_values('Vendor Name').reset_index(drop=True)
        else:
            vendor_summary_df = pd.DataFrame(columns=['Vendor Name', 'Total Taxable Value', 'Total TDS'])

        # --- SAVING LOGIC ---
        if custom_filename and custom_filename.strip():
            clean_name = custom_filename.strip()
            output_filename = f"{clean_name} TDS Working.xlsx"
        else:
            output_filename = f"Processed_TDS_Working.xlsx"

        output_full_path = os.path.join(output_folder, output_filename)

        with pd.ExcelWriter(output_full_path, engine='openpyxl') as writer:
            detailed_df.to_excel(writer, sheet_name='TDS Working', index=False)
            summary_start_row = len(detailed_df) + 4
            if not summary_df.empty:
                summary_df.to_excel(writer, sheet_name='TDS Working', index=False, startrow=summary_start_row)
            vendor_summary_df.to_excel(writer, sheet_name='Vendor Summary', index=False)

        # --- POST-PROCESSING (formatting) ---
        import openpyxl
        wb = openpyxl.load_workbook(output_full_path)
        ws = wb['TDS Working']

        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(ws.max_column)}{len(detailed_df) + 1}"
        ws.freeze_panes = "A2"
        style_header_row(ws, len(detailed_df.columns))
        autofit_columns(ws, detailed_df)

        numeric_cols = {detailed_df.columns.get_loc(c) + 1 for c in
                         ['Total', 'Tax Deducted at Source', 'Total After TDS Deduction', 'Checking for rates',
                          'Difference', 'Interest'] if c in detailed_df.columns}
        style_data_rows(ws, 2, len(detailed_df) + 1, 1, len(detailed_df.columns), numeric_cols)

        if not summary_df.empty:
            style_header_row(ws, len(summary_df.columns), row=summary_start_row + 1)
            summary_numeric_cols = {summary_df.columns.get_loc(c) + 1 for c in
                                     ['Total Tax Deducted', 'Total Interest'] if c in summary_df.columns}
            style_data_rows(ws, summary_start_row + 2, summary_start_row + 1 + len(summary_df),
                             1, len(summary_df.columns), summary_numeric_cols)

        vendor_ws = wb['Vendor Summary']
        if not vendor_summary_df.empty:
            vendor_ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(vendor_ws.max_column)}{len(vendor_summary_df) + 1}"
        vendor_ws.freeze_panes = "A2"
        style_header_row(vendor_ws, len(vendor_summary_df.columns))
        autofit_columns(vendor_ws, vendor_summary_df)
        vendor_numeric_cols = {vendor_summary_df.columns.get_loc(c) + 1 for c in
                                ['Total Taxable Value', 'Total TDS'] if c in vendor_summary_df.columns}
        style_data_rows(vendor_ws, 2, len(vendor_summary_df) + 1, 1, len(vendor_summary_df.columns), vendor_numeric_cols)

        wb.save(output_full_path)

        # Clean up input files
        for fp in file_paths:
            if os.path.exists(fp):
                os.remove(fp)

        return {
            "success": True,
            "message": f"Processed {len(file_paths)} file(s) successfully.",
            "download_url": f"/api/download/{output_filename}",
            "filename": output_filename,
            "summary_data": summary_df.to_dict(orient='records') if not summary_df.empty else []
        }

    except Exception as e:
        return {"success": False, "error": str(e)}
