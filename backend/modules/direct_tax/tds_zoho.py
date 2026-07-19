import pandas as pd
import os
from modules.excel_styles import style_header_row, style_data_rows, autofit_columns
from modules.direct_tax.tds_section_mapping import lookup_new_section as _lookup_new_section


def process_tds_zoho(file_paths, output_folder, custom_filename=None):
    """
    Processes multiple Zoho TDS export files (one per old section, e.g.
    194C.xlsx, 194J.xlsx -- the filename becomes the Old Section label) into
    a single TDS Working report with New Section/Section Code added per the
    Income Tax Act 2025 mapping above, plus a Vendor Summary sheet.
    """
    try:
        all_data = []

        # Helper to find header row in messy Excel files
        def find_header_row(file_path):
            try:
                df_sample = pd.read_excel(file_path, header=None, nrows=10)
                for i, row in df_sample.iterrows():
                    if row.astype(str).str.contains('Transaction#', case=False, na=False).any():
                        return i
                return 0
            except:
                return 0

        # 1. Read All Files
        for fp in file_paths:
            try:
                filename = os.path.basename(fp)
                section_name = os.path.splitext(filename)[0]

                if fp.endswith('.csv'):
                    df = pd.read_csv(fp, skip_blank_lines=True)
                else:
                    header_row = find_header_row(fp)
                    df = pd.read_excel(fp, header=header_row)

                df.dropna(how='all', inplace=True)
                df['Old Section'] = section_name
                all_data.append(df)
            except Exception as e:
                return {"success": False, "error": f"Failed to read {filename}: {str(e)}"}

        if not all_data:
            return {"success": False, "error": "No valid data found."}

        combined_df = pd.concat(all_data, ignore_index=True)

        # Standardize Columns
        master_rename_map = {
            'pan': 'Permanent Account Number (PAN)',
            'total': 'Total',
            'amount': 'Total',
            'tax deducted at source': 'Tax Deducted at Source',
            'tds': 'Tax Deducted at Source',
            'rate at which deducted': 'Rate at which deducted',
            'tds rate': 'Rate at which deducted',
            'date': 'Date',
            'transaction#': 'Transaction#',
            'vendor': 'Vendor',
            'transaction type': 'Transaction Type'
        }
        combined_df.rename(columns=lambda c: master_rename_map.get(str(c).strip().lower(), str(c).strip()), inplace=True)

        # Calculations
        df = combined_df
        df['Tax Deducted at Source'] = pd.to_numeric(df['Tax Deducted at Source'], errors='coerce').fillna(0).round(0)
        df['Total'] = pd.to_numeric(df['Total'], errors='coerce').fillna(0)
        df['Rate at which deducted'] = pd.to_numeric(df['Rate at which deducted'], errors='coerce').fillna(0)
        df['Total After TDS Deduction'] = df['Total'] - df['Tax Deducted at Source']

        def get_pan_type(pan):
            pan_str = str(pan)
            if len(pan_str) >= 4 and pan_str[3].upper() == 'C': return 'Co.'
            return 'Non Co.'

        if 'Permanent Account Number (PAN)' in df.columns:
            df['Co./Non Co.'] = df['Permanent Account Number (PAN)'].apply(get_pan_type)
        else:
            df['Co./Non Co.'] = 'Unknown'

        df['Checking for rate'] = df['Total'] * (df['Rate at which deducted'] / 100)
        df['Difference'] = df['Tax Deducted at Source'] - df['Checking for rate']
        df['Remarks'] = df['Difference'].apply(lambda x: 'match' if abs(x) < 1 else 'Mismatch')

        # New Section / Section Code (Income Tax Act 2025 mapping)
        new_sections, section_codes = [], []
        for _, row in df.iterrows():
            new_sec, code = _lookup_new_section(row['Old Section'], row['Rate at which deducted'])
            new_sections.append(new_sec)
            section_codes.append(code)
        df['New Section'] = new_sections
        df['Section Code'] = section_codes

        # Date Logic
        if 'Date' in df.columns:
            df['Date_dt'] = pd.to_datetime(df['Date'], errors='coerce')
            valid_dates = df['Date_dt'].dropna()
            report_month_period = None
            if not valid_dates.empty:
                modes = valid_dates.dt.to_period('M').mode()
                if not modes.empty: report_month_period = modes[0]

            if report_month_period:
                df['Months Difference'] = df['Date_dt'].dt.to_period('M').apply(lambda x: (report_month_period - x).n if pd.notna(x) else 0)
                df['Interest'] = df.apply(lambda row: row['Tax Deducted at Source'] * 0.015 * (row['Months Difference'] + 1) if row['Months Difference'] > 0 else 0, axis=1).round(0)
            else:
                df['Interest'] = 0
            df['Date'] = df['Date_dt'].dt.strftime('%Y-%m-%d').fillna('')
        else:
            df['Interest'] = 0

        # Reorder (blank Challan placeholder columns removed -- handled by
        # the separate TDS Challan Mapper tool instead)
        final_column_order = [
            'Transaction#', 'Date', 'Vendor', 'Permanent Account Number (PAN)',
            'Transaction Type', 'Total', 'Tax Deducted at Source', 'Total After TDS Deduction',
            'Checking for rate', 'Difference', 'Remarks', 'Interest', 'Rate at which deducted',
            'Old Section', 'New Section', 'Section Code', 'Co./Non Co.',
        ]
        existing_cols = [c for c in final_column_order if c in df.columns]
        detailed_df = df[existing_cols].copy()

        # Section Summary
        if {'Old Section', 'New Section', 'Section Code', 'Co./Non Co.'}.issubset(detailed_df.columns):
            for col in ['Old Section', 'New Section', 'Section Code', 'Co./Non Co.']:
                detailed_df[col] = detailed_df[col].astype(str).str.strip().fillna('N/A')
            summary_df = detailed_df.groupby(['Old Section', 'New Section', 'Section Code', 'Co./Non Co.']).agg(
                Total_Tax_Deducted=('Tax Deducted at Source', 'sum'),
                Total_Interest=('Interest', 'sum')
            ).reset_index()
            summary_df.rename(columns={'Total_Tax_Deducted': 'Total Tax Deducted', 'Total_Interest': 'Total Interest'}, inplace=True)
        else:
            summary_df = pd.DataFrame()

        # Vendor Summary
        if 'Vendor' in detailed_df.columns:
            vendor_summary_df = detailed_df.groupby('Vendor').agg(
                **{'Total Taxable Value': ('Total', 'sum'), 'Total TDS': ('Tax Deducted at Source', 'sum')}
            ).reset_index().rename(columns={'Vendor': 'Vendor Name'})
            vendor_summary_df = vendor_summary_df.sort_values('Vendor Name').reset_index(drop=True)
        else:
            vendor_summary_df = pd.DataFrame(columns=['Vendor Name', 'Total Taxable Value', 'Total TDS'])

        # --- SIMPLIFIED NAMING LOGIC (Matches Odoo) ---
        if custom_filename and str(custom_filename).strip():
            clean_name = str(custom_filename).strip()
            output_filename = f"{clean_name} TDS Working.xlsx"
        else:
            output_filename = "Zoho_Processed_TDS.xlsx"

        output_full_path = os.path.join(output_folder, output_filename)

        with pd.ExcelWriter(output_full_path, engine='openpyxl') as writer:
            detailed_df.to_excel(writer, sheet_name='TDS Working', index=False)
            summary_start_row = len(detailed_df) + 4
            if not summary_df.empty:
                summary_df.to_excel(writer, sheet_name='TDS Working', index=False, startrow=summary_start_row)
            vendor_summary_df.to_excel(writer, sheet_name='Vendor Summary', index=False)

        # --- POST-PROCESSING (formatting) ---
        import openpyxl
        wb = openpyxl.load_workbook(output_full_path)
        ws = wb['TDS Working']

        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(ws.max_column)}{len(detailed_df) + 1}"
        ws.freeze_panes = "A2"
        style_header_row(ws, len(detailed_df.columns))
        autofit_columns(ws, detailed_df)

        numeric_cols = {detailed_df.columns.get_loc(c) + 1 for c in
                         ['Total', 'Tax Deducted at Source', 'Total After TDS Deduction', 'Checking for rate',
                          'Difference', 'Interest'] if c in detailed_df.columns}
        style_data_rows(ws, 2, len(detailed_df) + 1, 1, len(detailed_df.columns), numeric_cols)

        if not summary_df.empty:
            style_header_row(ws, len(summary_df.columns), row=summary_start_row + 1)
            summary_numeric_cols = {summary_df.columns.get_loc(c) + 1 for c in
                                     ['Total Tax Deducted', 'Total Interest'] if c in summary_df.columns}
            style_data_rows(ws, summary_start_row + 2, summary_start_row + 1 + len(summary_df),
                             1, len(summary_df.columns), summary_numeric_cols)

        vendor_ws = wb['Vendor Summary']
        if not vendor_summary_df.empty:
            vendor_ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(vendor_ws.max_column)}{len(vendor_summary_df) + 1}"
        vendor_ws.freeze_panes = "A2"
        style_header_row(vendor_ws, len(vendor_summary_df.columns))
        autofit_columns(vendor_ws, vendor_summary_df)
        vendor_numeric_cols = {vendor_summary_df.columns.get_loc(c) + 1 for c in
                                ['Total Taxable Value', 'Total TDS'] if c in vendor_summary_df.columns}
        style_data_rows(vendor_ws, 2, len(vendor_summary_df) + 1, 1, len(vendor_summary_df.columns), vendor_numeric_cols)

        wb.save(output_full_path)

        # Cleanup input files
        for fp in file_paths:
            if os.path.exists(fp): os.remove(fp)

        return {
            "success": True,
            "message": "Processed successfully.",
            "download_url": f"/api/download/{output_filename}",
            "filename": output_filename,
            "summary_data": summary_df.to_dict(orient='records') if not summary_df.empty else []
        }

    except Exception as e:
        return {"success": False, "error": str(e)}
