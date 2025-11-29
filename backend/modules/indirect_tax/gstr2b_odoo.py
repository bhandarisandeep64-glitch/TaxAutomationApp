import pandas as pd
import numpy as np
import os
import re
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def clean_purchase_data(df, is_rcm=False):
    """
    Logic to clean Odoo Purchase data.
    """
    # Ensure required columns exist to avoid key errors
    required_cols = ['Account', 'Label', 'Date']
    for col in required_cols:
        if col not in df.columns: df[col] = ''

    # Handle Debit/Credit
    if 'Debit' not in df.columns: df['Debit'] = 0.0
    if 'Credit' not in df.columns: df['Credit'] = 0.0
    
    df['Debit'] = pd.to_numeric(df['Debit'], errors='coerce').fillna(0)
    df['Credit'] = pd.to_numeric(df['Credit'], errors='coerce').fillna(0)

    # --- TAX CALCULATION ---
    account_series = df['Account'].astype(str)
    is_igst = account_series.str.contains('igst', case=False, na=False)
    is_cgst = account_series.str.contains('cgst', case=False, na=False)
    is_sgst = account_series.str.contains('sgst', case=False, na=False)
    
    # Calculate Tax Amount
    if is_rcm:
        condition = df['Debit'] != 0
        df['Tax_Amount'] = np.where(condition, df['Debit'], df['Credit'])
    else:
        condition = df['Credit'] != 0
        df['Tax_Amount'] = np.where(condition, df['Credit'], df['Debit'])

    # Assign columns
    df['IGST'] = np.where(is_igst, df['Tax_Amount'], 0.0)
    df['CGST'] = np.where(is_cgst, df['Tax_Amount'], 0.0)
    df['SGST'] = np.where(is_sgst, df['Tax_Amount'], 0.0)

    # Auto-Fill SGST if missing
    mask_missing_sgst = (df['CGST'] != 0) & (df['SGST'] == 0)
    df.loc[mask_missing_sgst, 'SGST'] = df.loc[mask_missing_sgst, 'CGST']

    # Extract Rate
    if 'Label' in df.columns:
        df['Rate'] = df['Label'].astype(str).str.extract(r'(\d+\.?\d*)%').astype(float)
        df['Rate_Str'] = df['Rate'].map('{:.2f}%'.format, na_action='ignore')
    else:
        df['Rate'] = 0.0
        df['Rate_Str'] = '0.00%'
    
    # Calculate Total
    if 'Taxable Amt.' not in df.columns: df['Taxable Amt.'] = 0.0
    df['Taxable Amt.'] = pd.to_numeric(df['Taxable Amt.'], errors='coerce').fillna(0)
    
    df['Total'] = df['Taxable Amt.'] + df['IGST'] + df['CGST'] + df['SGST']
    
    # Date Format
    if 'Date' in df.columns:
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce').dt.strftime('%d-%m-%Y')

    # --- FINAL COLUMNS (Forced) ---
    final_columns = [
        'Partner', 'GSTIN', 'Date', 'Number', 'Reference', 'Account', 
        'Label', 'Rate_Str', 'Total', 'Taxable Amt.', 'IGST', 'CGST', 'SGST',
        'As per Portal', 'Difference', 'Remarks'
    ]
    
    # Force add columns if they don't exist
    for col in final_columns:
        if col not in df.columns:
            df[col] = '' 
            
    return df[final_columns]

def process_gstr2b_odoo(file_paths_dict, output_folder, custom_filename=None):
    """
    Wrapper: Accepts DICT of file paths.
    """
    try:
        processed_results = {
            'B2B as per Books': pd.DataFrame(),
            'V CN as per Books': pd.DataFrame(),
            'RCM as per Books': pd.DataFrame(),
            'V CN RCM as per Books': pd.DataFrame()
        }
        
        def load_slot(key):
            path = file_paths_dict.get(key)
            if path and os.path.exists(path):
                if path.endswith('.csv'): return pd.read_csv(path)
                return pd.read_excel(path)
            return None

        # A. Regular
        reg_cgst = load_slot('regular_cgst')
        reg_igst = load_slot('regular_igst')
        
        regular_dfs = [df for df in [reg_cgst, reg_igst] if df is not None]
        
        if regular_dfs:
            reg_df = pd.concat(regular_dfs, ignore_index=True)
            if 'Credit' in reg_df.columns:
                v_cn_mask = reg_df['Credit'] != 0
                processed_results['B2B as per Books'] = clean_purchase_data(reg_df[~v_cn_mask].copy(), is_rcm=False)
                processed_results['V CN as per Books'] = clean_purchase_data(reg_df[v_cn_mask].copy(), is_rcm=False)
            else:
                processed_results['B2B as per Books'] = clean_purchase_data(reg_df, is_rcm=False)

        # B. RCM
        rcm_cgst = load_slot('rcm_cgst')
        rcm_igst = load_slot('rcm_igst')
        
        rcm_dfs = [df for df in [rcm_cgst, rcm_igst] if df is not None]

        if rcm_dfs:
            rcm_df = pd.concat(rcm_dfs, ignore_index=True)
            if 'Debit' in rcm_df.columns:
                rcm_cn_mask = rcm_df['Debit'] != 0
                
                rcm_cn_data = rcm_df[rcm_cn_mask].copy()
                if 'Taxable Amt.' in rcm_cn_data.columns: 
                    rcm_cn_data['Taxable Amt.'] = rcm_cn_data['Taxable Amt.'] * -1
                
                processed_results['RCM as per Books'] = clean_purchase_data(rcm_df[~rcm_cn_mask].copy(), is_rcm=True)
                processed_results['V CN RCM as per Books'] = clean_purchase_data(rcm_cn_data, is_rcm=True)
            else:
                processed_results['RCM as per Books'] = clean_purchase_data(rcm_df, is_rcm=True)

        # Check data
        has_data = any(not df.empty for df in processed_results.values())
        if not has_data:
             return {"success": False, "error": "No valid data found in uploaded files."}

        # 4. Summary (Dashboard Only)
        summary_rows = []
        for category, df in processed_results.items():
            if not df.empty:
                summary_rows.append({
                    'Category': category,
                    'Taxable': df['Taxable Amt.'].sum(),
                    'IGST': df['IGST'].sum(),
                    'CGST': df['CGST'].sum(),
                    'SGST': df['SGST'].sum()
                })
        
        summary_df = pd.DataFrame(summary_rows)
        if not summary_df.empty:
            for col in ['Taxable', 'IGST', 'CGST', 'SGST']:
                summary_df[col] = summary_df[col].round(2)

        # 5. Save to Excel (With Formatting)
        if custom_filename and str(custom_filename).strip():
            clean_name = re.sub(r'[\\/*?:"<>|]', '-', str(custom_filename).strip())
            output_filename = f"{clean_name} Purchase Reco.xlsx"
        else:
            output_filename = "GSTR2B_Odoo_Processed.xlsx"

        output_full_path = os.path.join(output_folder, output_filename)

        # Write Sheets
        with pd.ExcelWriter(output_full_path, engine='openpyxl') as writer:
            for sheet_name, df in processed_results.items():
                if not df.empty:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Apply Formatting
        wb = openpyxl.load_workbook(output_full_path)
        sum_cols = ['Total', 'Taxable Amt.', 'IGST', 'CGST', 'SGST']

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            max_row = ws.max_row
            
            if max_row < 2: continue

            # AutoFilter
            ws.auto_filter.ref = ws.dimensions

            # Grand Total Row - Immediately after data (no gap)
            total_row_idx = max_row + 1
            ws.cell(row=total_row_idx, column=1).value = "GRAND TOTAL"
            ws.cell(row=total_row_idx, column=1).font = Font(bold=True)

            # Add Subtotals
            headers = [cell.value for cell in ws[1]]
            for col_name in sum_cols:
                if col_name in headers:
                    col_idx = headers.index(col_name) + 1
                    col_letter = get_column_letter(col_idx)
                    formula = f"=SUBTOTAL(9, {col_letter}2:{col_letter}{max_row})"
                    
                    cell = ws.cell(row=total_row_idx, column=col_idx)
                    cell.value = formula
                    cell.font = Font(bold=True)
                    cell.number_format = '#,##0.00'

        wb.save(output_full_path)

        return {
            "success": True,
            "message": "Purchase Registers Processed.",
            "download_url": f"/api/download/{output_filename}",
            "filename": output_filename,
            "summary_data": summary_df.to_dict(orient='records')
        }

    except Exception as e:
        return {"success": False, "error": str(e)}