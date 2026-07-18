import pandas as pd
import numpy as np
import re
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# ==========================================
#  HSN B2B / HSN B2C parser (current Odoo export format)
# ==========================================
#
# Odoo now exports exactly two files -- "HSN B2B" and "HSN B2C" -- from the
# account.move.line ("Journal Item") view. One row per line-item within an
# invoice (an invoice with several HSN codes/products gets one row per
# HSN/SAC code, all sharing the same invoice Number), rather than the old
# per-tax-account ledger export.
#
# Columns actually used: GSTR Section, Partner, GSTIN, Date, HSN/SAC Code,
# Number, Taxes (e.g. "18% GST S" / "18% IGST S" / "5% GST S"), Taxable Amt.
# (an invoice-level total, repeated across that invoice's lines), Credit,
# Debit.
#
# Verified against real exports (both files the firm shared) and a live
# client run cross-checked against Odoo's own official GSTR-1 report:
#   - Which file is B2B vs B2C has to come from the upload slot, not the
#     data -- the B2C file's own GSTR Section column is always "B2CS", even
#     for its credit notes, so it can't tell CDNR apart on its own.
#   - Credit-note detection: the row's own GSTR Section text ("CDNR
#     Regular") is the primary, reliable signal for the B2B file; Debit
#     being populated is used as a fallback for the B2C file (whose GSTR
#     Section never distinguishes CDNR) and belt-and-braces for the B2B
#     file. GSTR Section wins when the two disagree -- found one real
#     credit note posted with its amount in Credit instead of Debit, which
#     a Debit-only check misclassified as a regular sale.
#   - Credit/Debit is the TAXABLE VALUE for that specific line (not the tax
#     amount) -- confirmed against a real multi-line invoice where the line
#     values summed to the invoice's own "Taxable Amt." total, not to 18%
#     of it. Tax is computed here, not read from the file.
#   - "Taxes" gives the rate and type as plain text: "IGST S" -> the whole
#     tax is IGST; "GST S" -> the tax splits evenly into CGST/SGST.
#   - An invoice can mix rates across its lines (e.g. some products at 18%,
#     others at 5%), so tax is computed per LINE using that line's own rate,
#     then summed per invoice -- not one rate applied to the invoice total.
#   - Rounding happens ONCE, after summing the unrounded per-line tax across
#     an invoice (or HSN code) -- rounding per line first and summing
#     afterwards was found to let independent roundings drift apart across
#     many line splits, by more than a rupee on some invoices.
#   - Credit notes reduce the total -- they're stored as positive
#     magnitudes in their own "B2B CDNR"/"B2C CDNR" categories (matching the
#     real filed 3B working paper's own TOTAL = B2B - B2B CDNR + B2C - B2C
#     CDNR convention), but shown as NEGATIVE in the invoice-level detail
#     table and netted (not added) into the HSN-wise summary and the GRAND
#     TOTAL row, so a plain SUM/SUBTOTAL over the detail table nets
#     correctly instead of double-counting them as additional sales.
#   - The per-line sum can occasionally fall short of the file's own
#     "Taxable Amt." column (observed on one real invoice, off by ~10k) --
#     likely a line posted to an account outside this export. The per-line
#     sum is used as the authoritative figure (it's what the tax
#     calculation is actually built from); the file's own total is kept
#     alongside as a reference column, and any material mismatch is logged
#     rather than silently dropped.

RATE_RE = re.compile(r'(\d+(?:\.\d+)?)\s*%')

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TOTAL_FONT = Font(bold=True)
THIN_BORDER = Border(bottom=Side(style='thin', color="BFBFBF"))
NEGATIVE_FONT = Font(color="C00000")
NUMBER_FORMAT = '#,##0.00;[Red](#,##0.00)'


def _read_file(fp):
    if str(fp).lower().endswith('.csv'):
        return pd.read_csv(fp)
    return pd.read_excel(fp)


def _extract_rate(taxes_str):
    m = RATE_RE.search(str(taxes_str or ''))
    return float(m.group(1)) if m else 0.0


def _is_igst(taxes_str):
    return 'igst' in str(taxes_str or '').lower()


def _split_and_round(igst_raw, intrastate_raw):
    """Rounds a combined-tax figure exactly once, after summing unrounded
    line-level amounts across an invoice or HSN code -- rounding per line
    first and summing afterwards was found to accumulate drift of several
    paise (occasionally over a rupee) on invoices with many line splits.
    cgst/sgst always sum exactly to the rounded intrastate total, differing
    from each other by at most 1 paisa, and only when that total itself is
    an odd number of paise -- the same unavoidable split GST portals
    themselves show."""
    igst = igst_raw.round(2)
    cgst = (intrastate_raw / 2).round(2)
    sgst = intrastate_raw.round(2) - cgst
    return igst, cgst, sgst


def _parse_hsn_file(fp, base_category):
    """Returns a per-line DataFrame from one HSN B2B/B2C export, tagged with
    Nature (base_category, or base_category + ' CDNR' for credit notes) and
    the UNROUNDED per-line tax split into an IGST-attributable amount and an
    intrastate (CGST+SGST combined) amount -- rounding happens later, once,
    at the invoice/HSN aggregation level (see _split_and_round)."""
    df = _read_file(fp)

    for col in ['GSTIN', 'Partner', 'Number', 'Taxes', 'HSN/SAC Code', 'Date']:
        if col not in df.columns:
            df[col] = ''
    for col in ['Taxable Amt.', 'Credit', 'Debit']:
        if col not in df.columns:
            df[col] = 0.0
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    section = df['GSTR Section'].astype(str).str.lower()
    is_cn = section.str.contains('cdnr', na=False) | (df['Debit'] != 0)
    df['Nature'] = np.where(is_cn, f"{base_category} CDNR", base_category)
    # Per-line taxable value -- whichever of Debit/Credit is actually
    # populated for that row (only one ever is), independent of the Nature
    # tag above so a Debit/Credit swap can't also corrupt the amount.
    df['Line_Taxable'] = np.where(df['Debit'] != 0, df['Debit'], df['Credit'])
    df['Rate'] = df['Taxes'].apply(_extract_rate)
    df['Is_IGST'] = df['Taxes'].apply(_is_igst)

    line_tax = df['Line_Taxable'] * df['Rate'] / 100
    df['Line_IGST_Raw'] = np.where(df['Is_IGST'], line_tax, 0.0)
    df['Line_Intrastate_Raw'] = np.where(df['Is_IGST'], 0.0, line_tax)

    return df


def _invoice_level(per_line_df):
    """Collapses the per-line rows down to one row per invoice Number,
    summing the per-line taxable/tax figures (correct even when an invoice
    mixes tax rates across its lines) -- the file's own invoice-level
    'Taxable Amt.' is kept alongside as a reference/reconciliation column,
    not used for the calculation itself."""
    grouped = per_line_df.groupby('Number', as_index=False).agg(
        Nature=('Nature', 'first'), Partner=('Partner', 'first'),
        GSTIN=('GSTIN', 'first'), Date=('Date', 'first'),
        **{'HSN/SAC Code': ('HSN/SAC Code', 'first')},
        **{'Taxable Amt. (Odoo Reference)': ('Taxable Amt.', 'first')},
        **{'Taxable Amt.': ('Line_Taxable', 'sum')},
        IGST_Raw=('Line_IGST_Raw', 'sum'), Intrastate_Raw=('Line_Intrastate_Raw', 'sum'),
    )
    grouped['IGST'], grouped['CGST'], grouped['SGST'] = _split_and_round(
        grouped['IGST_Raw'], grouped['Intrastate_Raw']
    )
    return grouped.drop(columns=['IGST_Raw', 'Intrastate_Raw'])


def compute_gstr1_data(file_paths):
    """file_paths: dict with optional 'file_b2b'/'file_b2c' keys pointing at
    Odoo's HSN B2B / HSN B2C Journal Item exports. At least one is required.

    Returns (final_df, summary, hsn_summary_df):
      - final_df: one row per invoice, with IGST/CGST/SGST computed line by
        line and summed (so mixed-rate invoices come out correct). Credit
        notes are shown as NEGATIVE here (reducing the total on a plain
        SUM), even though they're a positive magnitude in `summary`.
      - summary: groupby('Nature') totals in the {Category, Taxable, IGST,
        CGST, SGST} shape gstr3b_engine.py's compute_gstr1_buckets expects
        (positive magnitudes for every category, credit notes included),
        plus a GRAND TOTAL row that nets credit notes out
        (B2B - B2B CDNR + B2C - B2C CDNR). This is the authoritative source
        for GSTR-1/3B totals.
      - hsn_summary_df: HSN-wise net taxable/tax totals (Table 12 style),
        netting credit notes against regular invoices per HSN code -- a
        useful new breakdown this format makes possible, but a best-effort
        one (see module docstring on the occasional per-line shortfall).
    """
    per_line_frames = []
    for key, base_category in (('file_b2b', 'B2B'), ('file_b2c', 'B2C')):
        fp = file_paths.get(key) if file_paths else None
        if fp:
            per_line_frames.append(_parse_hsn_file(fp, base_category))

    if not per_line_frames:
        return None, None, None

    per_line_df = pd.concat(per_line_frames, ignore_index=True)

    invoice_df = _invoice_level(per_line_df)
    for col in ['Taxable Amt.', 'IGST', 'CGST', 'SGST', 'Taxable Amt. (Odoo Reference)']:
        invoice_df[col] = invoice_df[col].round(2)
    invoice_df['Invoice Total'] = (
        invoice_df['Taxable Amt.'] + invoice_df['IGST'] + invoice_df['CGST'] + invoice_df['SGST']
    ).round(2)

    mismatch = (invoice_df['Taxable Amt.'] - invoice_df['Taxable Amt. (Odoo Reference)']).abs()
    flagged = invoice_df[mismatch > 1.0]
    if not flagged.empty:
        print(
            f"GSTR-1 note: {len(flagged)} invoice(s) where the computed taxable value "
            f"(summed from line items) differs from Odoo's own invoice total by more than "
            f"Rs 1 -- likely a line posted outside this export. Invoice numbers: "
            f"{list(flagged['Number'])}"
        )

    # Category summary FIRST, from the still-unsigned (positive-magnitude)
    # invoice_df -- this is what gstr3b_engine.compute_gstr1_buckets reads,
    # and it expects every category (including the CDNR ones) as a positive
    # magnitude, matching the real filed 3B working paper's own convention.
    summary = invoice_df.groupby('Nature').agg(
        Taxable=('Taxable Amt.', 'sum'), IGST=('IGST', 'sum'),
        CGST=('CGST', 'sum'), SGST=('SGST', 'sum'),
    ).reset_index().rename(columns={'Nature': 'Category'})

    cat = summary.set_index('Category')
    def _get(name, col):
        return float(cat.loc[name, col]) if name in cat.index else 0.0
    total_values = {
        col: _get('B2B', col) - _get('B2B CDNR', col) + _get('B2C', col) - _get('B2C CDNR', col)
        for col in ['Taxable', 'IGST', 'CGST', 'SGST']
    }
    total_values['Category'] = 'GRAND TOTAL'
    total_row = pd.DataFrame([total_values])
    summary = pd.concat([summary, total_row], ignore_index=True)
    for col in ['Taxable', 'IGST', 'CGST', 'SGST']:
        summary[col] = summary[col].round(2)

    # Detail table: flip sign on credit-note rows so a plain SUM/SUBTOTAL
    # over this sheet nets them instead of adding them as extra sales --
    # this only affects display; the category math above already used the
    # unsigned values.
    final_df = invoice_df.copy()
    is_cdnr_row = final_df['Nature'].str.contains('CDNR')
    for col in ['Taxable Amt.', 'IGST', 'CGST', 'SGST', 'Invoice Total']:
        final_df.loc[is_cdnr_row, col] = -final_df.loc[is_cdnr_row, col].abs()

    final_cols = ['Nature', 'Partner', 'GSTIN', 'Date', 'Number', 'HSN/SAC Code',
                  'Invoice Total', 'Taxable Amt.', 'IGST', 'CGST', 'SGST',
                  'Taxable Amt. (Odoo Reference)']
    final_df = final_df[[c for c in final_cols if c in final_df.columns]].copy()

    is_cdnr = per_line_df['Nature'].str.contains('CDNR')
    sign = np.where(is_cdnr, -1, 1)
    hsn_grouped = per_line_df.assign(
        Signed_Taxable=per_line_df['Line_Taxable'] * sign,
        Signed_IGST_Raw=per_line_df['Line_IGST_Raw'] * sign,
        Signed_Intrastate_Raw=per_line_df['Line_Intrastate_Raw'] * sign,
    ).groupby('HSN/SAC Code', as_index=False).agg(
        **{'Taxable Value': ('Signed_Taxable', 'sum')},
        IGST_Raw=('Signed_IGST_Raw', 'sum'), Intrastate_Raw=('Signed_Intrastate_Raw', 'sum'),
    )
    hsn_grouped['IGST'], hsn_grouped['CGST'], hsn_grouped['SGST'] = _split_and_round(
        hsn_grouped['IGST_Raw'], hsn_grouped['Intrastate_Raw']
    )
    hsn_grouped['Taxable Value'] = hsn_grouped['Taxable Value'].round(2)
    hsn_grouped['Total Tax'] = hsn_grouped['IGST'] + hsn_grouped['CGST'] + hsn_grouped['SGST']
    hsn_summary_df = hsn_grouped.drop(columns=['IGST_Raw', 'Intrastate_Raw']) \
        .sort_values('HSN/SAC Code').reset_index(drop=True)

    return final_df, summary, hsn_summary_df


def _style_header_row(ws, n_cols, row=1):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _style_data_rows(ws, first_row, last_row, first_col, last_col, numeric_cols):
    for r in range(first_row, last_row + 1):
        for c in range(first_col, last_col + 1):
            cell = ws.cell(row=r, column=c)
            if c in numeric_cols:
                cell.number_format = NUMBER_FORMAT
                if isinstance(cell.value, (int, float)) and cell.value < 0:
                    cell.font = NEGATIVE_FONT


def _autofit_columns(ws, df, min_width=10, max_width=40):
    for i, col in enumerate(df.columns, start=1):
        sample = df[col].astype(str).head(200)
        width = max([len(str(col))] + [len(v) for v in sample]) + 2
        ws.column_dimensions[get_column_letter(i)].width = max(min_width, min(width, max_width))


def process_gstr1_odoo(file_paths, output_folder, custom_filename=None):
    """
    Main Wrapper:
    1. Reads the HSN B2B / HSN B2C files.
    2. Computes per-invoice GSTR-1 data and category/HSN summaries.
    3. Generates a formatted, professional-looking Excel with Subtotals,
       AutoFilter, and an HSN Summary sheet.
    """
    try:
        final_df, summary, hsn_summary_df = compute_gstr1_data(file_paths)
        if final_df is None:
            return {"success": False, "error": "Could not process any valid data. Upload at least one of the HSN B2B / HSN B2C files."}

        if custom_filename and str(custom_filename).strip():
            clean_name = re.sub(r'[\\/*?:"<>|]', '-', str(custom_filename).strip())
            output_filename = f"{clean_name} GSTR1.xlsx"
        else:
            output_filename = "GSTR1_Report.xlsx"

        output_full_path = os.path.join(output_folder, output_filename)

        with pd.ExcelWriter(output_full_path, engine='openpyxl') as writer:
            final_df.to_excel(writer, sheet_name='GSTR1 Data', index=False)

            last_row = len(final_df) + 2
            summary_start_row = last_row + 4
            summary.to_excel(writer, sheet_name='GSTR1 Data', index=False, startrow=summary_start_row)

            hsn_summary_df.to_excel(writer, sheet_name='HSN Summary', index=False)

        # --- POST-PROCESSING (formatting) ---
        import openpyxl
        wb = openpyxl.load_workbook(output_full_path)
        ws = wb['GSTR1 Data']

        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{len(final_df) + 1}"
        ws.freeze_panes = "A2"
        _style_header_row(ws, len(final_df.columns))
        _autofit_columns(ws, final_df)

        numeric_cols = {final_df.columns.get_loc(c) + 1 for c in ['Invoice Total', 'Taxable Amt.', 'IGST', 'CGST', 'SGST', 'Taxable Amt. (Odoo Reference)'] if c in final_df.columns}
        _style_data_rows(ws, 2, len(final_df) + 1, 1, len(final_df.columns), numeric_cols)

        ws.cell(row=last_row, column=1).value = "GRAND TOTAL"
        for c in range(1, len(final_df.columns) + 1):
            ws.cell(row=last_row, column=c).font = TOTAL_FONT
            ws.cell(row=last_row, column=c).fill = TOTAL_FILL
            ws.cell(row=last_row, column=c).border = THIN_BORDER

        for col_name in ['Taxable Amt.', 'IGST', 'CGST', 'SGST']:
            if col_name in final_df.columns:
                col_idx = final_df.columns.get_loc(col_name) + 1
                col_letter = get_column_letter(col_idx)
                ws.cell(row=last_row, column=col_idx).value = f"=SUBTOTAL(9, {col_letter}2:{col_letter}{len(final_df)+1})"
                ws.cell(row=last_row, column=col_idx).number_format = NUMBER_FORMAT

        # Style the category summary block below the detail table
        _style_header_row(ws, len(summary.columns), row=summary_start_row + 1)
        summary_last_row = summary_start_row + len(summary) + 1
        grand_total_row = summary_start_row + len(summary) + 1
        for c in range(1, len(summary.columns) + 1):
            for r in range(summary_start_row + 2, summary_last_row + 1):
                cell = ws.cell(row=r, column=c)
                if c > 1:
                    cell.number_format = NUMBER_FORMAT
            gt_cell = ws.cell(row=grand_total_row, column=c)
            gt_cell.font = TOTAL_FONT
            gt_cell.fill = TOTAL_FILL

        hsn_ws = wb['HSN Summary']
        if len(hsn_summary_df) > 0:
            hsn_ws.auto_filter.ref = f"A1:{get_column_letter(hsn_ws.max_column)}{len(hsn_summary_df) + 1}"
        hsn_ws.freeze_panes = "A2"
        _style_header_row(hsn_ws, len(hsn_summary_df.columns))
        _autofit_columns(hsn_ws, hsn_summary_df)
        hsn_numeric_cols = {hsn_summary_df.columns.get_loc(c) + 1 for c in ['Taxable Value', 'IGST', 'CGST', 'SGST', 'Total Tax'] if c in hsn_summary_df.columns}
        _style_data_rows(hsn_ws, 2, len(hsn_summary_df) + 1, 1, len(hsn_summary_df.columns), hsn_numeric_cols)

        wb.save(output_full_path)

        # Cleanup
        for fp in file_paths.values():
            if fp and os.path.exists(fp): os.remove(fp)

        return {
            "success": True,
            "message": "GSTR-1 Report Generated.",
            "download_url": f"/api/download/{output_filename}",
            "filename": output_filename,
            "summary_data": summary.to_dict(orient='records')
        }

    except Exception as e:
        return {"success": False, "error": str(e)}
