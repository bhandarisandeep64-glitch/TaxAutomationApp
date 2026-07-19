"""
Regression test for gstr2b_reco_zoho_engine.py.

No real GSTR-2B/Zoho sample files exist in this repo, so this builds fake
ones in-memory that match the column shapes the engine's cleaning functions
expect, and checks the specific gaps that were fixed:
  A) a genuine amount mismatch on an otherwise identically-identified
     invoice (same invoice number + GSTIN) is now flagged "Mismatch"
     instead of silently appearing as two disconnected "missing" entries
  B) a normal exact match still works (regression safety for the new pass)
  C) previous-period marking (month was previously never threaded through
     from the route into the engine at all -- reco_dt was hardcoded None)
  D) Section 16(4) time-barred flag
  E) ITC reference sheets are preserved (not dropped) but not merged into
     reconciliation
  F) the new "Mismatches" discrepancy sheet picks up case A

Run directly: python tests/test_gstr2b_reco_zoho_engine.py
(also pytest-discoverable once pytest is added to the project)
"""
import os
import sys
from io import BytesIO

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import pandas as pd

from modules.indirect_tax.gstr2b_reco_zoho_engine import generate_reco_report_zoho


def _build_portal_file():
    b2b_cols = ['GSTIN of supplier', 'Trade/Legal name', 'Invoice number', 'Invoice Date',
                'Invoice Value (₹)', 'Place of supply', 'Supply Attract Reverse Charge', 'Rate (%)',
                'Taxable Value (₹)', 'Integrated Tax (₹)', 'Central Tax (₹)', 'State/UT Tax (₹)', 'Cess Amount (₹)']

    rows = [
        # A: same invoice + GSTIN as books, but a genuine ₹5000 discrepancy -> should be "Mismatch"
        ['27AAAAA0000A1Z5', 'Test Vendor A', 'INV-001', '05-12-2025', 17700, 'Maharashtra', 'No', 18, 15000.00, 2700.00, 0, 0, 0],
        # B: normal exact match, regression check
        ['27BBBBB0000B1Z5', 'Test Vendor B', 'INV-002', '06-12-2025', 11800, 'Maharashtra', 'No', 18, 10000.00, 1800.00, 0, 0, 0],
        # C: previous period, no books match
        ['27CCCCC0000C1Z5', 'Test Vendor C', 'OLD-001', '15-10-2025', 5900, 'Maharashtra', 'No', 18, 5000.00, 900.00, 0, 0, 0],
        # D: time-barred (FY 2022-23, Section 16(4) deadline was 30-Nov-2023)
        ['27DDDDD0000D1Z5', 'Test Vendor D', 'TB-001', '10-05-2022', 5900, 'Maharashtra', 'No', 18, 5000.00, 900.00, 0, 0, 0],
        # filler rows so the sheet clears the "B2B" conditional row-count check (needs >6 rows)
        ['27FFFFF0000F1Z5', 'Filler Vendor 1', 'FILL-001', '02-12-2025', 1180, 'Maharashtra', 'No', 18, 1000.00, 180.00, 0, 0, 0],
        ['27FFFFF0000F1Z5', 'Filler Vendor 1', 'FILL-002', '03-12-2025', 1180, 'Maharashtra', 'No', 18, 1000.00, 180.00, 0, 0, 0],
        ['27FFFFF0000F1Z5', 'Filler Vendor 1', 'FILL-003', '04-12-2025', 1180, 'Maharashtra', 'No', 18, 1000.00, 180.00, 0, 0, 0],
    ]
    b2b_df = pd.DataFrame(rows, columns=b2b_cols)

    itc_rows = [
        ['27AAAAA0000A1Z5', 'Test Vendor A', 'INV-001', '05-12-2025', 17700, 'Maharashtra', 'No', 18, 15000.00, 2700.00, 0, 0, 0],
        ['27ZZZZZ0000Z1Z5', 'Blocked Vendor Z', 'BLK-001', '06-12-2025', 1180, 'Maharashtra', 'No', 18, 1000.00, 180.00, 0, 0, 0],
    ]
    itc_df = pd.DataFrame(itc_rows, columns=b2b_cols)

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='xlsxwriter') as writer:
        b2b_df.to_excel(writer, sheet_name='B2B', header=True, index=False)
        itc_df.to_excel(writer, sheet_name='ITC Available', header=True, index=False)
    buf.seek(0)
    return buf


def _build_zoho_file():
    # clean_zoho_data reads with header=1, i.e. row 0 is a title row and
    # row 1 (Excel row 2) is the real header -- matches Zoho Books' export
    # format of one title row above the column headers.
    cols = ['GSTIN', 'Vendor Name', 'Invoice Number', 'Invoice Date', 'Taxable Value', 'IGST', 'CGST', 'SGST']
    rows = [
        # A: same invoice + GSTIN as portal, but ₹5000 higher -- genuine mismatch
        ['27AAAAA0000A1Z5', 'Test Vendor A', 'INV-001', '05-12-2025', 20000.00, 3600.00, 0, 0],
        # B: exact match
        ['27BBBBB0000B1Z5', 'Test Vendor B', 'INV-002', '06-12-2025', 10000.00, 1800.00, 0, 0],
        # filler matches
        ['27FFFFF0000F1Z5', 'Filler Vendor 1', 'FILL-001', '02-12-2025', 1000.00, 180.00, 0, 0],
        ['27FFFFF0000F1Z5', 'Filler Vendor 1', 'FILL-002', '03-12-2025', 1000.00, 180.00, 0, 0],
        ['27FFFFF0000F1Z5', 'Filler Vendor 1', 'FILL-003', '04-12-2025', 1000.00, 180.00, 0, 0],
    ]
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='xlsxwriter') as writer:
        # Row 0: title row (skipped by header=1), Row 1: real header, Row 2+: data
        title_and_data = pd.DataFrame([["Zoho Books B2B Export"] + [''] * (len(cols) - 1)] + [cols] + rows)
        title_and_data.to_excel(writer, sheet_name='B2B', header=False, index=False)
    buf.seek(0)
    return buf


def _build_portal_file_real_shape():
    """Models the *actual* raw GSTR-2B portal export shape (confirmed
    against a real download), which is different from _build_portal_file()
    above in the one way that mattered for a real bug: 'ITC Availability'
    is already a genuine per-invoice column sitting right on the B2B sheet
    itself (GSTN puts it there directly, values like 'Yes'/'No' per row),
    and the "ITC Available"/"ITC not available"/etc. sheets are NOT
    invoice-level lists at all -- they're the portal's own aggregate
    GSTR-3B-table-wise summary (Heading/GSTR-3B table/tax columns only, no
    GSTIN or Invoice Number anywhere). An earlier version of
    compute_reco_data_zoho assumed the opposite (eligibility merged in from
    those aggregate sheets) and, finding no per-invoice data there,
    overwrote the real 'ITC Availability' column with blanks -- silently
    zeroing every ITC bucket on real client files. This fixture exists so
    that regression can't quietly reappear."""
    b2b_cols = ['GSTIN of supplier', 'Trade/Legal name', 'Invoice number', 'Invoice Date',
                'Invoice Value (₹)', 'Place of supply', 'Supply Attract Reverse Charge', 'Rate (%)',
                'Taxable Value (₹)', 'Integrated Tax (₹)', 'Central Tax (₹)', 'State/UT Tax (₹)', 'Cess Amount (₹)',
                'ITC Availability']
    rows = [
        # needs >6 raw rows (header + 6 data rows) to clear clean_portal_data's
        # "B2B" conditional row-count check -- see _build_portal_file() above.
        ['27AAAAA0000A1Z5', 'Test Vendor A', 'INV-001', '05-12-2025', 11800, 'Maharashtra', 'No', 18, 10000.00, 1800.00, 0, 0, 0, 'Yes'],
        ['27FFFFF0000F1Z5', 'Filler Vendor 1', 'FILL-001', '02-12-2025', 1180, 'Maharashtra', 'No', 18, 1000.00, 180.00, 0, 0, 0, 'Yes'],
        ['27FFFFF0000F1Z5', 'Filler Vendor 1', 'FILL-002', '03-12-2025', 1180, 'Maharashtra', 'No', 18, 1000.00, 180.00, 0, 0, 0, 'Yes'],
        ['27FFFFF0000F1Z5', 'Filler Vendor 1', 'FILL-003', '04-12-2025', 1180, 'Maharashtra', 'No', 18, 1000.00, 180.00, 0, 0, 0, 'Yes'],
        ['27FFFFF0000F1Z5', 'Filler Vendor 1', 'FILL-004', '05-12-2025', 1180, 'Maharashtra', 'No', 18, 1000.00, 180.00, 0, 0, 0, 'Yes'],
        ['27FFFFF0000F1Z5', 'Filler Vendor 1', 'FILL-005', '06-12-2025', 1180, 'Maharashtra', 'No', 18, 1000.00, 180.00, 0, 0, 0, 'Yes'],
    ]
    b2b_df = pd.DataFrame(rows, columns=b2b_cols)

    # Real shape of the "ITC Available" reference sheet: aggregate totals
    # by GSTR-3B heading, no GSTIN/Invoice Number columns at all.
    itc_summary_df = pd.DataFrame([
        ['S.no.', 'Heading', 'GSTR-3B table', 'Integrated Tax  (₹)', 'Central Tax (₹)', 'State/UT Tax (₹)', 'Cess  (₹)'],
        ['I', 'All other ITC - Supplies from registered persons other than reverse charge', '4(A)(5)', 1800.0, 540.0, 0, 0],
    ])

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='xlsxwriter') as writer:
        b2b_df.to_excel(writer, sheet_name='B2B', header=True, index=False)
        itc_summary_df.to_excel(writer, sheet_name='ITC Available', header=False, index=False)
    buf.seek(0)
    return buf


def test_itc_availability_survives_real_shaped_portal_file():
    """The actual bug: with a real-shaped portal file (native per-invoice
    'ITC Availability' on B2B, aggregate-only "ITC Available" reference
    sheet), the B2B sheet's real 'Yes' values must survive
    compute_reco_data_zoho untouched -- not get overwritten with blanks by
    a merge step that was solving a problem that doesn't exist."""
    from modules.indirect_tax.gstr2b_reco_zoho_engine import compute_reco_data_zoho

    portal_file = _build_portal_file_real_shape()
    zoho_file = _build_zoho_file()

    processed_portal, _, _ = compute_reco_data_zoho(portal_file, zoho_file, '2025-12')
    b2b = processed_portal['B2B']

    assert 'ITC Availability' in b2b.columns, "ITC Availability column missing entirely"
    assert (b2b['ITC Availability'] == 'Yes').all(), \
        f"ITC Availability got wiped out: {b2b['ITC Availability'].tolist()}"

    print("ITC Availability survives a real-shaped portal file (not wiped by a phantom merge) OK")


def _get_row(df, inv):
    match = df[df['Invoice Number'] == inv]
    assert len(match) == 1, f"Expected exactly 1 row for invoice {inv!r}, found {len(match)}"
    return match.iloc[0]


def test_gstr2b_reco_zoho_engine():
    portal_file = _build_portal_file()
    zoho_file = _build_zoho_file()

    manual_inputs = {
        'sales': {'taxable': 0, 'igst': 0, 'cgst': 0, 'sgst': 0},
        'opening': {'igst': 0, 'cgst': 0, 'sgst': 0},
    }

    output = generate_reco_report_zoho(portal_file, zoho_file, manual_inputs, month_str="2025-12")
    all_sheets = pd.read_excel(output, sheet_name=None)

    b2b_portal = next((df for name, df in all_sheets.items() if name.startswith('B2B') and '(Portal)' in name), None)
    assert b2b_portal is not None, f"Could not find the B2B (Portal) sheet in output. Got: {list(all_sheets.keys())}"

    # A: genuine amount mismatch on identically-identified invoice -> "Mismatch", not "Not in Books"
    a = _get_row(b2b_portal, 'INV-001')
    assert a['Remarks'] == 'Mismatch', f"Test A failed: got {a['Remarks']!r}"
    assert abs(a['Difference'] - (-5000.0)) < 0.01, f"Test A difference wrong: got {a['Difference']!r}"

    # B: normal exact match still works
    b = _get_row(b2b_portal, 'INV-002')
    assert b['Remarks'] == 'Match', f"Test B failed: got {b['Remarks']!r}"

    # C: previous-period invoice with no books match
    c = _get_row(b2b_portal, 'OLD-001')
    assert c['Remarks'] == 'Previous Period Inv', f"Test C failed: got {c['Remarks']!r}"

    # D: Section 16(4) time-barred flag
    d = _get_row(b2b_portal, 'TB-001')
    assert str(d['ITC Time-Barred']).startswith('YES'), f"Test D failed: got {d['ITC Time-Barred']!r}"

    # E: ITC reference sheet preserved but not merged into reconciliation
    ref_sheets = [n for n in all_sheets if 'ITC Available' in n]
    assert len(ref_sheets) == 1 and '(Reference)' in ref_sheets[0], f"Test E failed: {ref_sheets}"

    # F: the new Mismatches discrepancy sheet picks up case A
    mismatches = all_sheets.get('Mismatches')
    assert mismatches is not None, "Test F failed: Mismatches sheet missing"
    assert (mismatches['Invoice Number'] == 'INV-001').any(), "Test F failed: INV-001 not in Mismatches sheet"

    print("ALL TESTS PASSED")


def test_master_dashboard_formulas_reference_correct_rows():
    """Real bug found on a live client file: several Master Dashboard
    formulas used hardcoded row numbers that were each off by one -- "5.
    Current Month ITC" pointed at a blank spacer row instead of "(C) Net ITC
    Available" (always showed 0), and both "NET PAYABLE IN CASH" and
    "BALANCE CREDIT C/F" summed an offset range that excluded "Paid by
    IGST" entirely and included a blank row instead -- silently dropping the
    IGST portion of the offset from the final cash-payable figure. Checks
    the actual formula strings, not just that a value happens to come out
    right for one dataset (which is exactly how this slipped through
    before)."""
    import openpyxl

    portal_file = _build_portal_file()
    zoho_file = _build_zoho_file()
    manual_inputs = {
        'sales': {'igst': 5000.0, 'cgst': 1000.0, 'sgst': 1000.0},
        'opening': {'igst': 0.0, 'cgst': 0.0, 'sgst': 0.0},
    }

    output = generate_reco_report_zoho(portal_file, zoho_file, manual_inputs, '2025-12')
    wb = openpyxl.load_workbook(output, data_only=False)
    ws = wb['Master Dashboard']

    def formula(desc_substring, col_letter):
        for row in range(1, ws.max_row + 1):
            if desc_substring in str(ws.cell(row=row, column=1).value or ''):
                return ws.cell(row=row, column=3).value if col_letter == 'C' else ws.cell(row=row, column=5).value
        return None

    # "5. Current Month ITC" must reference row 12 ("(C) Net ITC Available"),
    # not row 13 (a blank spacer).
    f = formula('Current Month ITC', 'C')
    assert f == '=C12', f"Current Month ITC formula wrong: {f}"

    # "NET PAYABLE IN CASH" must sum rows 20:22 ("Paid by IGST" through
    # "Paid by SGST"), not 21:23 (which excludes IGST and includes a blank).
    f = formula('NET PAYABLE', 'C')
    assert f == '=MAX(0, C5-SUM(C20:C22))', f"NET PAYABLE formula wrong: {f}"

    # "BALANCE CREDIT C/F" must reference row 17 ("6. Total Credit
    # Available"), not row 18 (blank), with the same corrected offset range.
    f = formula('BALANCE CREDIT', 'C')
    assert f == '=MAX(0, C17-SUM(C20:C22))', f"BALANCE CREDIT formula wrong: {f}"

    print("Master Dashboard formulas reference the correct rows: OK")


if __name__ == '__main__':
    test_gstr2b_reco_zoho_engine()
    test_master_dashboard_formulas_reference_correct_rows()
    test_itc_availability_survives_real_shaped_portal_file()
